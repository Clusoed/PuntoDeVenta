"""
Módulo de Importación/Exportación de Inventario desde Excel
Permite cargar inventario masivamente y descargar plantillas.
"""
import os
import sys
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from datetime import datetime
import customtkinter as ctk
from tkinter import filedialog
from CTkMessagebox import CTkMessagebox

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

# Agregar path para imports
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from database import (
    get_categorias, crear_categoria, get_producto_por_codigo,
    crear_producto, actualizar_producto, generar_codigo_producto
)
from utils.theme import (
    BG_PRINCIPAL, BG_SECUNDARIO, BORDER_COLOR, 
    TEXT_PRIMARY, TEXT_SECONDARY, ACCENT_PRIMARY
)


# Columnas de la plantilla Excel
COLUMNAS_PLANTILLA = [
    ("Código", "Código único del producto (ej: PROD001, SKU-123). Dejar vacío para generar automático."),
    ("Nombre", "Nombre del producto (OBLIGATORIO)"),
    ("Precio USD", "Precio de venta en USD (OBLIGATORIO)"),
    ("Categoría", "Nombre de la categoría (se crea si no existe)"),
    ("Marca", "Marca del producto (opcional)"),
    ("Unidad", "Unidad de medida: Unidad, Kg, Lt, etc. (default: Unidad)"),
    ("Costo USD", "Costo de compra en USD (opcional, default: 0)"),
    ("% Ganancia", "Porcentaje de ganancia (opcional, default: 30)"),
    ("Stock", "Stock actual (opcional, default: 0)"),
    ("Stock Mínimo", "Stock mínimo para alertas (opcional, default: 5)"),
]


def generar_plantilla_excel(ruta: str = None) -> str:
    """
    Genera una plantilla Excel con el formato correcto para importar inventario.
    Retorna la ruta del archivo generado.
    """
    if Workbook is None:
        raise ImportError("openpyxl no está instalado")
    
    # Si no se especifica ruta, usar el escritorio
    if not ruta:
        desktop = Path.home() / "Desktop"
        if not desktop.exists():
            desktop = Path.home() / "Escritorio"
        if not desktop.exists():
            desktop = Path.home()
        ruta = str(desktop / "Plantilla_Inventario.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir encabezados
    for col, (nombre, descripcion) in enumerate(COLUMNAS_PLANTILLA, 1):
        cell = ws.cell(row=1, column=col, value=nombre)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Ajustar ancho de columna
        ws.column_dimensions[get_column_letter(col)].width = max(15, len(nombre) + 5)
    
    # Agregar fila de ejemplo
    ejemplo = [
        "PROD001",          # Código
        "Arroz Premium",    # Nombre
        2.50,               # Precio USD
        "Alimentos",        # Categoría
        "Marca ABC",        # Marca
        "Kg",               # Unidad
        1.80,               # Costo USD
        30,                 # % Ganancia
        100,                # Stock
        10                  # Stock Mínimo
    ]
    
    for col, valor in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col, value=valor)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Agregar segunda fila de ejemplo
    ejemplo2 = [
        "",                 # Código (vacío para auto-generar)
        "Aceite Vegetal",   # Nombre
        4.00,               # Precio USD
        "Alimentos",        # Categoría
        "",                 # Marca (sin marca)
        "Lt",               # Unidad
        3.00,               # Costo USD
        25,                 # % Ganancia
        50,                 # Stock
        5                   # Stock Mínimo
    ]
    
    for col, valor in enumerate(ejemplo2, 1):
        cell = ws.cell(row=3, column=col, value=valor)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Crear hoja de instrucciones
    ws_instrucciones = wb.create_sheet("Instrucciones")
    instrucciones = [
        "INSTRUCCIONES PARA IMPORTAR INVENTARIO",
        "",
        "1. Complete los datos en la hoja 'Inventario'",
        "2. Los campos OBLIGATORIOS son: Nombre y Precio USD",
        "3. Si deja el Código vacío, se generará uno automático",
        "4. Si la Categoría no existe, se creará automáticamente",
        "5. Elimine las filas de ejemplo antes de importar",
        "",
        "DESCRIPCIÓN DE COLUMNAS:",
    ]
    
    for i, texto in enumerate(instrucciones, 1):
        ws_instrucciones.cell(row=i, column=1, value=texto)
    
    fila = len(instrucciones) + 1
    for nombre, descripcion in COLUMNAS_PLANTILLA:
        ws_instrucciones.cell(row=fila, column=1, value=f"• {nombre}:")
        ws_instrucciones.cell(row=fila, column=2, value=descripcion)
        fila += 1
    
    ws_instrucciones.column_dimensions['A'].width = 20
    ws_instrucciones.column_dimensions['B'].width = 60
    
    # Guardar
    wb.save(ruta)
    return ruta


def validar_fila(fila: Dict, fila_num: int, categorias_existentes: Dict) -> Tuple[bool, str, Dict]:
    """
    Valida una fila de datos del Excel.
    Retorna: (es_valido, mensaje_error, datos_limpios)
    """
    errores = []
    datos = {}
    
    # Nombre (obligatorio)
    nombre = str(fila.get("Nombre", "")).strip()
    if not nombre:
        errores.append("El nombre es obligatorio")
    else:
        datos["nombre"] = nombre
    
    # Precio USD (obligatorio)
    try:
        precio = float(fila.get("Precio USD", 0) or 0)
        if precio <= 0:
            errores.append("El precio debe ser mayor a 0")
        else:
            datos["precio_usd"] = precio
    except (ValueError, TypeError):
        errores.append("Precio USD inválido")
    
    # Código (opcional, se genera si está vacío)
    codigo = str(fila.get("Código", "")).strip()
    if codigo:
        datos["codigo"] = codigo
    else:
        datos["codigo"] = None  # Se generará automáticamente
    
    # Categoría (opcional)
    categoria = str(fila.get("Categoría", "")).strip()
    if categoria:
        if categoria.lower() in categorias_existentes:
            datos["categoria_id"] = categorias_existentes[categoria.lower()]
        else:
            # Marcar para crear categoría nueva
            datos["nueva_categoria"] = categoria
    else:
        datos["categoria_id"] = None
    
    # Marca (opcional)
    marca = str(fila.get("Marca", "")).strip()
    datos["marca"] = marca if marca else None
    
    # Unidad (opcional, default: Unidad)
    unidad = str(fila.get("Unidad", "")).strip()
    datos["unidad_medida"] = unidad if unidad else "Unidad"
    
    # Costo USD (opcional)
    try:
        costo = float(fila.get("Costo USD", 0) or 0)
        datos["costo_usd"] = max(0, costo)
    except (ValueError, TypeError):
        datos["costo_usd"] = 0
    
    # % Ganancia (opcional)
    try:
        ganancia = float(fila.get("% Ganancia", 30) or 30)
        datos["porcentaje_ganancia"] = max(0, ganancia)
    except (ValueError, TypeError):
        datos["porcentaje_ganancia"] = 30
    
    # Stock (opcional)
    try:
        stock = int(float(fila.get("Stock", 0) or 0))
        datos["stock_actual"] = max(0, stock)
    except (ValueError, TypeError):
        datos["stock_actual"] = 0
    
    # Stock Mínimo (opcional)
    try:
        stock_min = int(float(fila.get("Stock Mínimo", 5) or 5))
        datos["stock_minimo"] = max(0, stock_min)
    except (ValueError, TypeError):
        datos["stock_minimo"] = 5
    
    if errores:
        return False, f"Fila {fila_num}: " + ", ".join(errores), {}
    
    return True, "", datos


def importar_inventario(ruta: str, actualizar_existentes: bool = True) -> Tuple[int, int, int, List[str]]:
    """
    Importa inventario desde un archivo Excel.
    
    Args:
        ruta: Ruta del archivo Excel
        actualizar_existentes: Si True, actualiza productos con código existente
        
    Retorna: (creados, actualizados, errores, lista_mensajes_error)
    """
    if Workbook is None:
        raise ImportError("openpyxl no está instalado")
    
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active
    
    # Obtener categorías existentes
    categorias = get_categorias()
    categorias_dict = {cat['nombre'].lower(): cat['id'] for cat in categorias}
    
    # Leer encabezados
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip())
    
    # Procesar filas
    creados = 0
    actualizados = 0
    errores = 0
    mensajes_error = []
    
    for fila_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        # Saltar filas vacías
        if not any(row):
            continue
        
        # Crear diccionario de la fila
        fila_dict = {}
        for i, valor in enumerate(row):
            if i < len(headers):
                fila_dict[headers[i]] = valor
        
        # Validar fila
        es_valido, mensaje, datos = validar_fila(fila_dict, fila_num, categorias_dict)
        
        if not es_valido:
            errores += 1
            mensajes_error.append(mensaje)
            continue
        
        try:
            # Crear categoría si es nueva
            if "nueva_categoria" in datos:
                nueva_cat_id = crear_categoria(datos["nueva_categoria"])
                categorias_dict[datos["nueva_categoria"].lower()] = nueva_cat_id
                datos["categoria_id"] = nueva_cat_id
                del datos["nueva_categoria"]
            
            # Generar código si está vacío
            if datos["codigo"] is None:
                datos["codigo"] = generar_codigo_producto()
            
            # Verificar si el producto ya existe
            producto_existente = get_producto_por_codigo(datos["codigo"])
            
            if producto_existente:
                if actualizar_existentes:
                    # Actualizar producto existente
                    actualizar_producto(
                        producto_id=producto_existente['id'],
                        codigo=datos["codigo"],
                        nombre=datos["nombre"],
                        precio_usd=datos["precio_usd"],
                        categoria_id=datos.get("categoria_id"),
                        marca=datos.get("marca"),
                        unidad_medida=datos.get("unidad_medida"),
                        costo_usd=datos.get("costo_usd"),
                        porcentaje_ganancia=datos.get("porcentaje_ganancia"),
                        stock_minimo=datos.get("stock_minimo")
                    )
                    actualizados += 1
                else:
                    errores += 1
                    mensajes_error.append(f"Fila {fila_num}: El código '{datos['codigo']}' ya existe")
            else:
                # Crear nuevo producto
                crear_producto(
                    codigo=datos["codigo"],
                    nombre=datos["nombre"],
                    precio_usd=datos["precio_usd"],
                    categoria_id=datos.get("categoria_id"),
                    marca=datos.get("marca"),
                    unidad_medida=datos.get("unidad_medida", "Unidad"),
                    costo_usd=datos.get("costo_usd", 0),
                    porcentaje_ganancia=datos.get("porcentaje_ganancia", 30),
                    stock_actual=datos.get("stock_actual", 0),
                    stock_minimo=datos.get("stock_minimo", 5)
                )
                creados += 1
                
        except Exception as e:
            errores += 1
            mensajes_error.append(f"Fila {fila_num}: Error al procesar - {str(e)}")
    
    wb.close()
    return creados, actualizados, errores, mensajes_error


class ImportDialog(ctk.CTkToplevel):
    """Diálogo para importar inventario desde Excel."""
    
    def __init__(self, parent, callback=None):
        super().__init__(parent)
        
        self.callback = callback
        self.archivo_seleccionado = None
        
        self.title("Importar Inventario desde Excel")
        self.geometry("550x400")
        self.resizable(False, False)
        
        # Configurar color de fondo
        self.configure(fg_color=BG_PRINCIPAL)
        
        # Centrar ventana
        self.update_idletasks()
        x = (self.winfo_screenwidth() - 550) // 2
        y = (self.winfo_screenheight() - 400) // 2
        self.geometry(f"550x400+{x}+{y}")
        
        # Modal
        self.transient(parent)
        self.grab_set()
        
        self._create_widgets()
        self.protocol("WM_DELETE_WINDOW", self._cerrar)
    
    def _create_widgets(self):
        # Frame principal
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        ctk.CTkLabel(
            main_frame,
            text="📥 Importar Inventario desde Excel",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=TEXT_PRIMARY
        ).pack(pady=(0, 15))
        
        # Instrucciones
        instrucciones = ctk.CTkLabel(
            main_frame,
            text="Seleccione un archivo Excel con los productos a importar.\n"
                 "Use el botón 'Descargar Plantilla' para obtener el formato correcto.",
            font=ctk.CTkFont(size=12),
            text_color=TEXT_SECONDARY,
            justify="center"
        )
        instrucciones.pack(pady=(0, 15))
        
        # Frame de selección de archivo
        file_frame = ctk.CTkFrame(main_frame, fg_color=BG_SECUNDARIO)
        file_frame.pack(fill="x", pady=10)
        
        self.file_label = ctk.CTkLabel(
            file_frame,
            text="Ningún archivo seleccionado",
            font=ctk.CTkFont(size=11),
            text_color=TEXT_SECONDARY
        )
        self.file_label.pack(pady=10, padx=10)
        
        ctk.CTkButton(
            file_frame,
            text="📂 Seleccionar Archivo",
            command=self._seleccionar_archivo,
            fg_color=ACCENT_PRIMARY,
            width=180
        ).pack(pady=(0, 10))
        
        # Checkbox para actualizar existentes
        self.actualizar_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            main_frame,
            text="Actualizar productos existentes (si el código ya existe)",
            variable=self.actualizar_var,
            text_color=TEXT_PRIMARY
        ).pack(pady=15)
        
        # Frame de resultados (oculto inicialmente)
        self.results_frame = ctk.CTkFrame(main_frame, fg_color=BG_SECUNDARIO)
        self.results_label = ctk.CTkLabel(
            self.results_frame,
            text="",
            font=ctk.CTkFont(size=12),
            text_color=TEXT_PRIMARY,
            justify="left"
        )
        self.results_label.pack(pady=10, padx=10)
        
        # Frame de botones
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(pady=20)
        
        ctk.CTkButton(
            button_frame,
            text="📄 Descargar Plantilla",
            command=self._descargar_plantilla,
            fg_color="gray50",
            width=150
        ).pack(side="left", padx=5)
        
        self.import_btn = ctk.CTkButton(
            button_frame,
            text="✅ Importar",
            command=self._importar,
            fg_color="#28a745",
            hover_color="#218838",
            width=150,
            state="disabled"
        )
        self.import_btn.pack(side="left", padx=5)
        
        ctk.CTkButton(
            button_frame,
            text="Cerrar",
            command=self._cerrar,
            fg_color="#dc3545",
            hover_color="#c82333",
            width=100
        ).pack(side="left", padx=5)
    
    def _seleccionar_archivo(self):
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            self.archivo_seleccionado = archivo
            nombre = os.path.basename(archivo)
            self.file_label.configure(text=f"📎 {nombre}", text_color=ACCENT_PRIMARY)
            self.import_btn.configure(state="normal")
    
    def _descargar_plantilla(self):
        try:
            ruta = filedialog.asksaveasfilename(
                title="Guardar plantilla",
                defaultextension=".xlsx",
                initialfile="Plantilla_Inventario.xlsx",
                filetypes=[("Archivo Excel", "*.xlsx")]
            )
            
            if ruta:
                generar_plantilla_excel(ruta)
                CTkMessagebox(
                    title="Plantilla Generada",
                    message=f"La plantilla se guardó en:\n{ruta}",
                    icon="check"
                )
        except Exception as e:
            CTkMessagebox(
                title="Error",
                message=f"Error al generar plantilla:\n{str(e)}",
                icon="cancel"
            )
    
    def _importar(self):
        if not self.archivo_seleccionado:
            return
        
        self.import_btn.configure(state="disabled", text="⏳ Importando...")
        self.update()
        
        try:
            creados, actualizados, errores, mensajes = importar_inventario(
                self.archivo_seleccionado,
                actualizar_existentes=self.actualizar_var.get()
            )
            
            # Mostrar resultados
            resultado = f"✅ Creados: {creados}\n"
            resultado += f"🔄 Actualizados: {actualizados}\n"
            resultado += f"❌ Errores: {errores}"
            
            self.results_label.configure(text=resultado)
            self.results_frame.pack(fill="x", pady=10)
            
            if mensajes:
                # Mostrar errores detallados
                errores_texto = "\n".join(mensajes[:10])  # Limitar a 10
                if len(mensajes) > 10:
                    errores_texto += f"\n... y {len(mensajes) - 10} errores más"
                
                CTkMessagebox(
                    title="Errores Encontrados",
                    message=errores_texto,
                    icon="warning"
                )
            else:
                CTkMessagebox(
                    title="Importación Completada",
                    message=f"Se procesaron {creados + actualizados} productos correctamente.",
                    icon="check"
                )
            
            # Callback para actualizar la vista
            if self.callback:
                self.callback()
                
        except Exception as e:
            CTkMessagebox(
                title="Error",
                message=f"Error al importar:\n{str(e)}",
                icon="cancel"
            )
        finally:
            self.import_btn.configure(state="normal", text="✅ Importar")
    
    def _cerrar(self):
        self.destroy()
