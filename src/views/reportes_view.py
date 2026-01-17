"""
Vista de Reportes del Sistema - Diseño Minimalista
Con funcionalidad de Exportación a Excel
"""
import customtkinter as ctk
from CTkMessagebox import CTkMessagebox
from datetime import datetime, timedelta
import sys
import os
from tkinter import filedialog
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from database import get_connection, get_tasa_actual, get_productos
from utils.currency import formato_usd, formato_bs
from utils.theme import BG_PRINCIPAL, BG_SECUNDARIO, BORDER_COLOR, TEXT_PRIMARY, TEXT_SECONDARY, ACCENT_PRIMARY, BG_HOVER, ACCENT_HOVER


class ReportesView(ctk.CTkFrame):
    """Vista de reportes del sistema."""
    
    def __init__(self, parent, app_controller):
        super().__init__(parent, fg_color=BG_PRINCIPAL)
        self.app = app_controller
        self.tasa = get_tasa_actual()
        
        self.setup_ui()
    
    def setup_ui(self):
        """Configura la interfaz de reportes."""
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        # === HEADER ===
        frame_header = ctk.CTkFrame(self, fg_color="transparent")
        frame_header.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        ctk.CTkLabel(
            frame_header,
            text="📈 REPORTES DEL SISTEMA",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(side="left")
        
        # Botones de exportación
        frame_export = ctk.CTkFrame(frame_header, fg_color="transparent")
        frame_export.pack(side="right")
        
        ctk.CTkButton(
            frame_export,
            text="📥 Exportar Ventas",
            fg_color="#28a745",
            hover_color="#218838",
            width=130,
            command=self.exportar_ventas
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            frame_export,
            text="📥 Exportar Compras",
            fg_color="#17a2b8",
            hover_color="#138496",
            width=130,
            command=self.exportar_compras
        ).pack(side="left", padx=3)
        
        ctk.CTkButton(
            frame_export,
            text="📥 Exportar Inventario",
            fg_color="#6c757d",
            hover_color="#545b62",
            width=140,
            command=self.exportar_inventario
        ).pack(side="left", padx=3)
        
        # === CONTENIDO ===
        frame_contenido = ctk.CTkFrame(self, fg_color=BG_SECUNDARIO, border_color=BORDER_COLOR, border_width=1)
        frame_contenido.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        frame_contenido.grid_columnconfigure((0, 1), weight=1)
        frame_contenido.grid_rowconfigure(1, weight=1)
        
        # === FILTROS ===
        frame_filtros = ctk.CTkFrame(frame_contenido, fg_color="transparent")
        frame_filtros.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=10)
        
        ctk.CTkLabel(frame_filtros, text="Período:").pack(side="left", padx=10)
        
        self.cmb_periodo = ctk.CTkComboBox(
            frame_filtros,
            values=["Hoy", "Esta Semana", "Este Mes", "Último Mes", "Todo"],
            width=150
        )
        self.cmb_periodo.pack(side="left", padx=5)
        self.cmb_periodo.set("Hoy")
        
        ctk.CTkButton(
            frame_filtros,
            text="🔄 Generar Reporte",
            fg_color=ACCENT_PRIMARY,
            hover_color=ACCENT_HOVER,
            command=self.generar_reporte
        ).pack(side="left", padx=20)
        
        # === TARJETAS DE RESUMEN ===
        frame_resumen = ctk.CTkFrame(frame_contenido, fg_color="transparent")
        frame_resumen.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
        frame_resumen.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        # Tarjeta Ventas Totales
        self.tarjeta_ventas = self.crear_tarjeta(
            frame_resumen, "💰 VENTAS TOTALES", "$ 0.00", "Bs 0.00", ACCENT_PRIMARY
        )
        self.tarjeta_ventas.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        
        # Tarjeta Número de Ventas
        self.tarjeta_cantidad = self.crear_tarjeta(
            frame_resumen, "🧾 TRANSACCIONES", "0", "facturas", "#3498db"
        )
        self.tarjeta_cantidad.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        # Tarjeta Promedio por Venta
        self.tarjeta_promedio = self.crear_tarjeta(
            frame_resumen, "📊 PROMEDIO/VENTA", "$ 0.00", "Bs 0.00", "#f39c12"
        )
        self.tarjeta_promedio.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        
        # Tarjeta Productos Vendidos
        self.tarjeta_productos = self.crear_tarjeta(
            frame_resumen, "📦 PRODUCTOS", "0", "unidades", "#9b59b6"
        )
        self.tarjeta_productos.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        
        # === TABLA DE VENTAS ===
        frame_tabla = ctk.CTkFrame(frame_contenido, fg_color="transparent")
        frame_tabla.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
        frame_tabla.grid_columnconfigure(0, weight=1)
        frame_tabla.grid_rowconfigure(1, weight=1)
        
        ctk.CTkLabel(
            frame_tabla,
            text="📋 DETALLE DE VENTAS",
            font=ctk.CTkFont(size=16, weight="bold")
        ).grid(row=0, column=0, sticky="w", padx=10, pady=10)
        
        self.frame_tabla_ventas = ctk.CTkScrollableFrame(frame_tabla, height=300)
        self.frame_tabla_ventas.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        for i in range(7):
            self.frame_tabla_ventas.grid_columnconfigure(i, weight=1)
        
        # Cargar datos iniciales
        self.generar_reporte()
    
    def crear_tarjeta(self, parent, titulo, valor1, valor2, color):
        """Crea una tarjeta de resumen con estilo minimalista."""
        tarjeta = ctk.CTkFrame(
            parent, 
            fg_color=BG_SECUNDARIO, 
            border_color=color,
            border_width=2,
            corner_radius=10
        )
        
        ctk.CTkLabel(
            tarjeta,
            text=titulo,
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=TEXT_SECONDARY
        ).pack(pady=(10, 5))
        
        lbl_valor1 = ctk.CTkLabel(
            tarjeta,
            text=valor1,
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=color
        )
        lbl_valor1.pack()
        
        lbl_valor2 = ctk.CTkLabel(
            tarjeta,
            text=valor2,
            font=ctk.CTkFont(size=14),
            text_color=TEXT_SECONDARY
        )
        lbl_valor2.pack(pady=(0, 10))
        
        tarjeta.lbl_valor1 = lbl_valor1
        tarjeta.lbl_valor2 = lbl_valor2
        
        return tarjeta
    
    def generar_reporte(self):
        """Genera el reporte según el período seleccionado."""
        periodo = self.cmb_periodo.get()
        
        # Calcular fechas
        hoy = datetime.now().date()
        
        if periodo == "Hoy":
            fecha_inicio = hoy
            # Agregar 1 día extra para capturar ventas con hora UTC
            fecha_fin = hoy + timedelta(days=1)
        elif periodo == "Esta Semana":
            fecha_inicio = hoy - timedelta(days=hoy.weekday())
            fecha_fin = hoy + timedelta(days=1)
        elif periodo == "Este Mes":
            fecha_inicio = hoy.replace(day=1)
            fecha_fin = hoy + timedelta(days=1)
        elif periodo == "Último Mes":
            primer_dia_mes = hoy.replace(day=1)
            fecha_fin = primer_dia_mes
            fecha_inicio = fecha_fin.replace(day=1) - timedelta(days=1)
            fecha_inicio = fecha_inicio.replace(day=1)
        else:  # Todo
            fecha_inicio = datetime(2020, 1, 1).date()
            fecha_fin = hoy + timedelta(days=1)
        
        # Obtener datos
        datos = self.obtener_datos_ventas(fecha_inicio, fecha_fin)
        
        # Actualizar tarjetas
        total_usd = datos['total_usd'] or 0
        total_bs = total_usd * self.tasa
        cantidad = datos['cantidad'] or 0
        promedio_usd = total_usd / cantidad if cantidad > 0 else 0
        promedio_bs = promedio_usd * self.tasa
        productos = datos['productos_vendidos'] or 0
        
        self.tarjeta_ventas.lbl_valor1.configure(text=formato_usd(total_usd))
        self.tarjeta_ventas.lbl_valor2.configure(text=formato_bs(total_bs))
        
        self.tarjeta_cantidad.lbl_valor1.configure(text=str(cantidad))
        
        self.tarjeta_promedio.lbl_valor1.configure(text=formato_usd(promedio_usd))
        self.tarjeta_promedio.lbl_valor2.configure(text=formato_bs(promedio_bs))
        
        self.tarjeta_productos.lbl_valor1.configure(text=str(productos))
        
        # Actualizar tabla
        self.mostrar_ventas(datos['ventas'])
    
    def obtener_datos_ventas(self, fecha_inicio, fecha_fin):
        """Obtiene datos de ventas del período."""
        conn = get_connection()
        cursor = conn.cursor()
        
        # Formatear fechas para SQLite
        fecha_ini_str = f"{fecha_inicio.isoformat()} 00:00:00"
        fecha_fin_str = f"{fecha_fin.isoformat()} 23:59:59"
        
        # Totales (incluir todas las ventas, sin filtro de estado por ahora)
        cursor.execute('''
            SELECT 
                COUNT(*) as cantidad,
                COALESCE(SUM(total_usd), 0) as total_usd
            FROM ventas
            WHERE fecha BETWEEN ? AND ?
        ''', (fecha_ini_str, fecha_fin_str))
        
        row = cursor.fetchone()
        cantidad = row['cantidad']
        total_usd = row['total_usd']
        
        # Productos vendidos
        cursor.execute('''
            SELECT COALESCE(SUM(dv.cantidad), 0) as total_productos
            FROM detalle_ventas dv
            JOIN ventas v ON dv.venta_id = v.id
            WHERE v.fecha BETWEEN ? AND ?
        ''', (fecha_ini_str, fecha_fin_str))
        
        productos = cursor.fetchone()['total_productos']
        
        # Lista de ventas
        cursor.execute('''
            SELECT numero_factura, fecha, total_usd, total_bs, forma_pago, 
                   tasa_cambio, COALESCE(estado, 'Completada') as estado
            FROM ventas
            WHERE fecha BETWEEN ? AND ?
            ORDER BY fecha DESC
            LIMIT 100
        ''', (fecha_ini_str, fecha_fin_str))
        
        ventas = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return {
            'cantidad': cantidad,
            'total_usd': total_usd,
            'productos_vendidos': productos,
            'ventas': ventas
        }
    
    def mostrar_ventas(self, ventas):
        """Muestra las ventas en la tabla."""
        # Limpiar tabla
        for widget in self.frame_tabla_ventas.winfo_children():
            widget.destroy()
        
        # Headers
        headers = ["Factura", "Fecha", "Total USD", "Total Bs", "Forma Pago", "Tasa", "Estado"]
        for i, header in enumerate(headers):
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text=header,
                font=ctk.CTkFont(size=12, weight="bold"),
                fg_color="#1a1a2e",
                padx=8,
                pady=6
            ).grid(row=0, column=i, sticky="ew", padx=1, pady=1)
        
        if not ventas:
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text="No hay ventas en este período",
                text_color="gray"
            ).grid(row=1, column=0, columnspan=7, pady=30)
            return
        
        for idx, venta in enumerate(ventas, start=1):
            bg = "#16213e" if idx % 2 == 0 else "transparent"
            
            # Factura
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text=venta['numero_factura'],
                fg_color=bg,
                padx=8, pady=4
            ).grid(row=idx, column=0, sticky="ew", padx=1)
            
            # Fecha
            fecha = venta['fecha'][:16] if venta['fecha'] else ""
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text=fecha,
                fg_color=bg,
                padx=8, pady=4
            ).grid(row=idx, column=1, sticky="ew", padx=1)
            
            # Total USD
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text=formato_usd(venta['total_usd']),
                fg_color=bg,
                text_color="#0ea5e9",
                padx=8, pady=4
            ).grid(row=idx, column=2, sticky="ew", padx=1)
            
            # Total Bs
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text=formato_bs(venta['total_bs']),
                fg_color=bg,
                text_color="#ffa500",
                padx=8, pady=4
            ).grid(row=idx, column=3, sticky="ew", padx=1)
            
            # Forma Pago
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text=venta['forma_pago'],
                fg_color=bg,
                padx=8, pady=4
            ).grid(row=idx, column=4, sticky="ew", padx=1)
            
            # Tasa
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text=f"Bs {venta['tasa_cambio']:,.2f}",
                fg_color=bg,
                padx=8, pady=4
            ).grid(row=idx, column=5, sticky="ew", padx=1)
            
            # Estado
            color_estado = "#28a745" if venta['estado'] == 'Completada' else "#dc3545"
            ctk.CTkLabel(
                self.frame_tabla_ventas,
                text=venta['estado'],
                fg_color=bg,
                text_color=color_estado,
                padx=8, pady=4
            ).grid(row=idx, column=6, sticky="ew", padx=1)
    
    # ==================== FUNCIONES DE EXPORTACIÓN ====================
    
    def _crear_estilo_excel(self):
        """Crea estilos reutilizables para Excel."""
        return {
            'header_font': Font(bold=True, color="FFFFFF", size=11),
            'header_fill': PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
            'header_alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def exportar_ventas(self):
        """Exporta todas las ventas a Excel."""
        if not OPENPYXL_AVAILABLE:
            CTkMessagebox(
                title="Error",
                message="La librería openpyxl no está instalada.\nInstálela con: pip install openpyxl",
                icon="cancel"
            )
            return
        
        # Obtener datos
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT 
                v.numero_factura,
                v.fecha,
                v.total_usd,
                v.total_bs,
                v.forma_pago,
                v.tasa_cambio,
                COALESCE(v.estado, 'Completada') as estado,
                v.observaciones
            FROM ventas v
            ORDER BY v.fecha DESC
        ''')
        ventas = cursor.fetchall()
        conn.close()
        
        if not ventas:
            CTkMessagebox(title="Aviso", message="No hay ventas para exportar.", icon="info")
            return
        
        # Seleccionar ubicación
        fecha_hoy = datetime.now().strftime("%Y%m%d")
        ruta = filedialog.asksaveasfilename(
            title="Guardar Ventas",
            defaultextension=".xlsx",
            initialfile=f"Ventas_{fecha_hoy}.xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")]
        )
        
        if not ruta:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ventas"
            estilos = self._crear_estilo_excel()
            
            # Headers
            headers = ["# Factura", "Fecha", "Total USD", "Total Bs", "Forma Pago", "Tasa", "Estado", "Observaciones"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = estilos['header_font']
                cell.fill = estilos['header_fill']
                cell.alignment = estilos['header_alignment']
                cell.border = estilos['border']
            
            # Datos
            for row, venta in enumerate(ventas, 2):
                ws.cell(row=row, column=1, value=venta['numero_factura'])
                ws.cell(row=row, column=2, value=venta['fecha'])
                ws.cell(row=row, column=3, value=venta['total_usd'])
                ws.cell(row=row, column=4, value=venta['total_bs'])
                ws.cell(row=row, column=5, value=venta['forma_pago'])
                ws.cell(row=row, column=6, value=venta['tasa_cambio'])
                ws.cell(row=row, column=7, value=venta['estado'])
                ws.cell(row=row, column=8, value=venta['observaciones'] or "")
            
            # Ajustar anchos
            for col in range(1, 9):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            wb.save(ruta)
            CTkMessagebox(
                title="Exportación Exitosa",
                message=f"Se exportaron {len(ventas)} ventas.\n\nArchivo: {Path(ruta).name}",
                icon="check"
            )
        except Exception as e:
            CTkMessagebox(title="Error", message=f"Error al exportar:\n{str(e)}", icon="cancel")
    
    def exportar_compras(self):
        """Exporta todas las compras a Excel."""
        if not OPENPYXL_AVAILABLE:
            CTkMessagebox(
                title="Error",
                message="La librería openpyxl no está instalada.\nInstálela con: pip install openpyxl",
                icon="cancel"
            )
            return
        
        # Obtener datos
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT 
                c.numero_factura,
                c.fecha,
                p.nombre as proveedor,
                c.total_usd,
                c.total_bs,
                c.forma_pago,
                c.tasa_cambio,
                COALESCE(c.estado, 'Completada') as estado,
                c.observaciones
            FROM compras c
            LEFT JOIN proveedores p ON c.proveedor_id = p.id
            ORDER BY c.fecha DESC
        ''')
        compras = cursor.fetchall()
        conn.close()
        
        if not compras:
            CTkMessagebox(title="Aviso", message="No hay compras para exportar.", icon="info")
            return
        
        # Seleccionar ubicación
        fecha_hoy = datetime.now().strftime("%Y%m%d")
        ruta = filedialog.asksaveasfilename(
            title="Guardar Compras",
            defaultextension=".xlsx",
            initialfile=f"Compras_{fecha_hoy}.xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")]
        )
        
        if not ruta:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Compras"
            estilos = self._crear_estilo_excel()
            estilos['header_fill'] = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
            
            # Headers
            headers = ["# Factura", "Fecha", "Proveedor", "Total USD", "Total Bs", "Forma Pago", "Tasa", "Estado", "Observaciones"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = estilos['header_font']
                cell.fill = estilos['header_fill']
                cell.alignment = estilos['header_alignment']
                cell.border = estilos['border']
            
            # Datos
            for row, compra in enumerate(compras, 2):
                ws.cell(row=row, column=1, value=compra['numero_factura'])
                ws.cell(row=row, column=2, value=compra['fecha'])
                ws.cell(row=row, column=3, value=compra['proveedor'] or "Sin proveedor")
                ws.cell(row=row, column=4, value=compra['total_usd'])
                ws.cell(row=row, column=5, value=compra['total_bs'])
                ws.cell(row=row, column=6, value=compra['forma_pago'])
                ws.cell(row=row, column=7, value=compra['tasa_cambio'])
                ws.cell(row=row, column=8, value=compra['estado'])
                ws.cell(row=row, column=9, value=compra['observaciones'] or "")
            
            # Ajustar anchos
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            wb.save(ruta)
            CTkMessagebox(
                title="Exportación Exitosa",
                message=f"Se exportaron {len(compras)} compras.\n\nArchivo: {Path(ruta).name}",
                icon="check"
            )
        except Exception as e:
            CTkMessagebox(title="Error", message=f"Error al exportar:\n{str(e)}", icon="cancel")
    
    def exportar_inventario(self):
        """Exporta el inventario actual a Excel."""
        if not OPENPYXL_AVAILABLE:
            CTkMessagebox(
                title="Error",
                message="La librería openpyxl no está instalada.\nInstálela con: pip install openpyxl",
                icon="cancel"
            )
            return
        
        # Obtener datos
        productos = get_productos()
        
        if not productos:
            CTkMessagebox(title="Aviso", message="No hay productos en inventario.", icon="info")
            return
        
        # Seleccionar ubicación
        fecha_hoy = datetime.now().strftime("%Y%m%d")
        ruta = filedialog.asksaveasfilename(
            title="Guardar Inventario",
            defaultextension=".xlsx",
            initialfile=f"Inventario_{fecha_hoy}.xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")]
        )
        
        if not ruta:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"
            estilos = self._crear_estilo_excel()
            estilos['header_fill'] = PatternFill(start_color="6c757d", end_color="6c757d", fill_type="solid")
            
            # Headers
            headers = ["Código", "Nombre", "Categoría", "Marca", "Unidad", "Precio USD", "Costo USD", 
                      "% Ganancia", "Stock Actual", "Stock Mínimo", "Estado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = estilos['header_font']
                cell.fill = estilos['header_fill']
                cell.alignment = estilos['header_alignment']
                cell.border = estilos['border']
            
            # Datos
            for row, prod in enumerate(productos, 2):
                ws.cell(row=row, column=1, value=prod.get('codigo', ''))
                ws.cell(row=row, column=2, value=prod.get('nombre', ''))
                ws.cell(row=row, column=3, value=prod.get('categoria_nombre', ''))
                ws.cell(row=row, column=4, value=prod.get('marca', ''))
                ws.cell(row=row, column=5, value=prod.get('unidad_medida', 'Unidad'))
                ws.cell(row=row, column=6, value=prod.get('precio_usd', 0))
                ws.cell(row=row, column=7, value=prod.get('costo_usd', 0))
                ws.cell(row=row, column=8, value=prod.get('porcentaje_ganancia', 30))
                ws.cell(row=row, column=9, value=prod.get('stock_actual', 0))
                ws.cell(row=row, column=10, value=prod.get('stock_minimo', 5))
                ws.cell(row=row, column=11, value="Activo" if prod.get('activo', 1) else "Inactivo")
            
            # Ajustar anchos
            anchos = [12, 30, 15, 15, 10, 12, 12, 12, 12, 12, 10]
            for col, ancho in enumerate(anchos, 1):
                ws.column_dimensions[get_column_letter(col)].width = ancho
            
            wb.save(ruta)
            CTkMessagebox(
                title="Exportación Exitosa",
                message=f"Se exportaron {len(productos)} productos.\n\nArchivo: {Path(ruta).name}",
                icon="check"
            )
        except Exception as e:
            CTkMessagebox(title="Error", message=f"Error al exportar:\n{str(e)}", icon="cancel")
